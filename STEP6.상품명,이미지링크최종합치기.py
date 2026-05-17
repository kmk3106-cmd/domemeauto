# -*- coding: utf-8 -*-
"""
엑셀 4개를 상품번호(도매매상품번호) 기준으로 매칭하여
1번 파일의 B열(쿠팡상품명), E열(상품이미지링크), G열(네이버카테고리ID)을 채워 넣는 스크립트.

- 1번: 2번_통합상품명.xlsx                      (A=상품번호, B=쿠팡상품명, E=이미지링크, G=네이버카테고리ID)
- 2번: 2번_통합상품명_keywords_쭌쭌쌤.xlsx     (A=상품번호, C=쿠팡상품명)
- 3번: 2번_domeme_links.xlsx                   (A=상품번호, E=상품이미지링크)
- 4번: 2번_통합상품명_카테고리매핑_결과.xlsx   (A=상품ID, C=네이버카테고리ID)

요구사항:
- 1번 B열 ← 2번의 C열(쿠팡상품명) 매칭하여 채움 (없으면 기존 B 유지)
- 1번 E열 ← 3번의 E열(이미지링크) 매칭하여 채움 (없으면 기존 E 유지)
- 1번 G열 ← 4번의 C열(네이버카테고리ID) 매칭하여 채움 (없으면 기존 G 유지)
- B열(상품명) 동일 중복 시 1,2,3 숫자 붙여 고유화 (전송 시 동일상품명 불가)

결과: 원본(1번)을 덮어쓰지 않고 "2번_통합상품명_최종.xlsx" 로 저장 (PermissionError 방지)

필요 패키지: pandas, openpyxl
pip install pandas openpyxl
"""

import os
import re
import subprocess
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict

import pandas as pd
from openpyxl import load_workbook

# 여기를 당신의 폴더로 지정하세요
BASE = Path(r"C:\Users\USER\Documents\국내위탁\마이박스\26년3월3주차\6번사업자")

# ===== 사용자 설정 (같은 폴더에 파일이 있다고 가정) =====
FILE1 = BASE / "6번_통합상품명.xlsx"  # 읽기 전용으로 사용 (업데이트 소스)
FILE2 = BASE / "6번_통합상품명_keywords_쭌쭌쌤.xlsx"  # A=상품번호, C=쿠팡상품명
FILE3 = BASE / "6번_domeme_links.xlsx"  # A=상품번호, E=상품이미지링크
FILE4 = BASE / "6번_통합상품명_카테고리매핑_결과.xlsx"  # A=상품ID, C=네이버카테고리ID
# 결과는 별도 파일로 저장 (원본 덮어쓰기 시 PermissionError 방지)
OUTPUT = BASE / "6번_통합상품명_최종.xlsx"
# True면 저장 후 Excel로 열었다가 그대로 저장 (스피드고전송기 등 업로드 호환용)
OPEN_AND_SAVE_IN_EXCEL = True

SHEET1 = 0
SHEET2 = 0
SHEET3 = 0
SHEET4 = 0


def norm_id(x: Any) -> str:
    """
    상품번호 정규화:
    - 공백 제거
    - float의 '.0' 제거
    - 숫자만 추출(정수형으로 쓰는 상품번호 가정)
    """
    if x is None:
        return ""
    s = str(x).strip()
    s = s.replace(",", "").replace(" ", "")
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    digits = re.findall(r"\d+", s)
    return "".join(digits) if digits else s


def excel_open_and_save(file_path: Path) -> bool:
    """Windows: Excel로 파일을 열었다가 그대로 저장 후 닫기 (호환용). 추가 패키지 없음."""
    if sys.platform != "win32":
        return False
    path_abs = file_path.resolve()
    if not path_abs.is_file():
        return False
    # 경로는 인자로 전달해 한글 경로 깨짐 방지 (VBS 본문은 ASCII만)
    vbs = """
On Error Resume Next
Dim path
path = WScript.Arguments(0)
Set objExcel = CreateObject("Excel.Application")
objExcel.Visible = False
objExcel.DisplayAlerts = False
Set wb = objExcel.Workbooks.Open(path)
If Not wb Is Nothing Then
    wb.Save
    wb.Close False
End If
objExcel.Quit
Set objExcel = Nothing
"""
    try:
        fd, vbs_path = tempfile.mkstemp(suffix=".vbs", text=True)
        try:
            os.write(fd, vbs.encode("ascii"))
            os.close(fd)
            subprocess.run(
                ["cscript", "//nologo", vbs_path, str(path_abs)],
                check=True, timeout=60,
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
            )
            return True
        finally:
            try:
                os.unlink(vbs_path)
            except OSError:
                pass
    except Exception:
        return False


def build_map_from_file(
    file_path: Path, sheet, id_col_idx: int, value_col_idx: int
) -> Dict[str, Any]:
    """엑셀에서 (id_col_idx)열을 키로, (value_col_idx)열을 값으로 하는 dict 생성"""
    df = pd.read_excel(file_path, sheet_name=sheet, header=0, dtype=str)
    ids = df.iloc[:, id_col_idx].map(norm_id)
    vals = df.iloc[:, value_col_idx]
    vals = vals.where(pd.notna(vals), None)
    mapping = {}
    for k, v in zip(ids, vals):
        if k:
            mapping[k] = v
    return mapping


def main():
    # 2번: A(0)=상품번호 → C(2)=쿠팡상품명
    name_map = build_map_from_file(FILE2, SHEET2, id_col_idx=0, value_col_idx=2)
    # 3번: A(0)=상품번호 → E(4)=이미지링크
    img_map = build_map_from_file(FILE3, SHEET3, id_col_idx=0, value_col_idx=4)
    # 4번: A(0)=상품ID → C(2)=네이버카테고리ID
    category_map = build_map_from_file(FILE4, SHEET4, id_col_idx=0, value_col_idx=2)

    # 1번 파일 로드 후 셀 위치로 직접 업데이트 (B, E, G 열)
    wb = load_workbook(FILE1, read_only=False)
    ws = wb[wb.sheetnames[SHEET1]]

    total_rows = 0
    name_updates = 0
    img_updates = 0
    category_updates = 0

    for row in ws.iter_rows(min_row=2):
        total_rows += 1
        cell_id = row[0]
        cell_name = row[1]
        cell_img = row[4]
        cell_category = row[6]

        pid = norm_id(cell_id.value)

        # B열: 2번 파일 C열
        new_name = name_map.get(pid, None)
        if new_name is not None and str(new_name).strip() != "":
            if str(cell_name.value or "").strip() != str(new_name).strip():
                cell_name.value = new_name
                name_updates += 1

        # E열: 3번 파일 E열
        new_img = img_map.get(pid, None)
        if new_img is not None and str(new_img).strip() != "":
            if str(cell_img.value or "").strip() != str(new_img).strip():
                cell_img.value = new_img
                img_updates += 1

        # G열: 4번 파일 C열
        new_category = category_map.get(pid, None)
        if new_category is not None and str(new_category).strip() != "":
            if str(cell_category.value or "").strip() != str(new_category).strip():
                cell_category.value = new_category
                category_updates += 1

    # B열(상품명) 동일 중복 검사 → 1,2,3 숫자 붙여 고유화 (전송 시 동일상품명 불가)
    name_total = defaultdict(int)
    for row in ws.iter_rows(min_row=2):
        val = str(row[1].value or "").strip()
        if val:
            name_total[val] += 1
    dup_names = {k for k, v in name_total.items() if v > 1}
    name_counter = defaultdict(int)
    dedup_count = 0
    for row in ws.iter_rows(min_row=2):
        cell_name = row[1]
        val = str(cell_name.value or "").strip()
        if not val or val not in dup_names:
            continue
        name_counter[val] += 1
        new_val = val + str(name_counter[val])
        cell_name.value = new_val
        dedup_count += 1
    if dedup_count > 0:
        print(f"B열(상품명) 중복 {len(dup_names)}종 {dedup_count}건 → 1,2,3 숫자 부여 완료")

    # 원본이 아닌 새 파일로 저장 → PermissionError 방지
    wb.save(OUTPUT)

    # Excel로 열었다가 그대로 저장 (스피드고전송기 등 업로드 호환용)
    if OPEN_AND_SAVE_IN_EXCEL and excel_open_and_save(OUTPUT):
        print("[Excel] 파일을 Excel로 열었다가 저장했습니다. (업로드 호환용)")

    print("------ 처리 결과 ------")
    print(f"총 행수(헤더 제외): {total_rows}")
    print(f"B열(쿠팡상품명) 갱신 건수: {name_updates}")
    if dedup_count > 0:
        print(f"B열(상품명) 중복 제거: {len(dup_names)}종 {dedup_count}건 → 1,2,3 숫자 부여")
    print(f"E열(이미지링크) 갱신 건수: {img_updates}")
    print(f"G열(네이버카테고리ID) 갱신 건수: {category_updates}")
    print(f"저장 파일: {OUTPUT}")


if __name__ == "__main__":
    main()
